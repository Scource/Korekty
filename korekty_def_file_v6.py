import xlsxwriter
from openpyxl import load_workbook, Workbook
import xlwings as xw
from datetime import datetime, timedelta, date
import xlrd


def volumen_add(column, MB_name):
    wolumen_sum = 0
    
    new_xlsx=open_xls_as_xlsx(MB_name[4])
    wb1=new_xlsx            
    ws=wb1.active
    day_vol=[]
    #energy_list=[]
    if ws['A29'].value != None:
        range_stop = 31
        for wolumen in range(6,range_stop):
            if wolumen == 8:
                cell_cord = column + str(wolumen)
                cell_cord_add = column + str(wolumen-1)
                day_vol.append(ws[cell_cord].value+ws[cell_cord_add].value)
                wolumen_sum += ws[cell_cord].value
            elif wolumen >= 10 or wolumen == 6 or wolumen == 7:
                cell_cord = column + str(wolumen-1)
                day_vol.append(ws[cell_cord].value)
                wolumen_sum += ws[cell_cord].value

    elif ws['A28'].value == None:
        range_stop = 30
        for wolumen in range(6,range_stop):
            if wolumen == 8:
                #cell_cord = column + str(wolumen)
                #cell_cord_add = column + str(wolumen-1)
                day_vol.append(0.0)
                wolumen_sum += 0.0
            elif wolumen == 6 or wolumen == 7:
                cell_cord = column + str(wolumen-1)
                day_vol.append(ws[cell_cord].value)
                wolumen_sum += ws[cell_cord].value
            else:
                cell_cord = column + str(wolumen-2)
                day_vol.append(ws[cell_cord].value)
                wolumen_sum += ws[cell_cord].value

    else:
        range_stop = 30
        for wolumen in range(6,range_stop):
            cell_cord = column + str(wolumen-1)
            day_vol.append(ws[cell_cord].value)
            wolumen_sum += ws[cell_cord].value

    
    return day_vol

def wire_volumen_sum(WIRE,POB_name):
    POB_P=[]
    POB_O=[]
    for MB_name in WIRE:
        if MB_name[1] == POB_name[1]:            
            POB_P_sum=volumen_add('A',MB_name)
            
            POB_P.append(POB_P_sum)
            
            POB_O_sum=volumen_add('C',MB_name)
            POB_O.append(POB_O_sum)
    return POB_P, POB_O
    
def import_WIRE_file(WIRE,POB_name, oddanie_column, pobor_column, oddanie_row, pobór_row, excel_file_name, plik_excel):
    cell_format=plik_excel.add_format({'font_size':'11', 'border':True, 'align':'center'})
    colOP=pobor_column
    colOO=oddanie_column
    for MB_name in WIRE:
        rowOP=pobór_row
        rowOO=oddanie_row
        
        if MB_name[1] == POB_name[1]:
            
            new_xlsx=open_xls_as_xlsx(MB_name[4])
            wb1=new_xlsx            
            ws=wb1.active
            O_O_sum = 0
            if ws['A29'].value != None:
                
                for O_P in range(6,31):
                    if O_P == 8:
                        excel_file_name.write_formula(rowOP,colOP, "='"+str(MB_name[3])+"DGMB'!B"+str(O_P)+"+'"+str(MB_name[3])+"DGMB'!B"+str(O_P+1),cell_format)
                        rowOP+=1
                    elif O_P >= 10 or O_P == 6 or O_P == 7:
                        excel_file_name.write_formula(rowOP,colOP, "='"+str(MB_name[3])+"DGMB'!B"+str(O_P),cell_format)
                        rowOP+=1
                for O_O in range(6,31):
                    cell_cord = 'C'+ str(O_O-1)
                    O_O_sum += ws[cell_cord].value
                    
                    if O_O == 8:
                        excel_file_name.write_formula(rowOO,colOO, "='"+str(MB_name[3])+"DGMB'!D"+str(O_O)+"+'"+str(MB_name[3])+"DGMB'!D"+str(O_O+1),cell_format)
                        rowOO+=1
                    elif O_O >= 10 or O_O == 6 or O_O == 7:
                        excel_file_name.write_formula(rowOO,colOO, "='"+str(MB_name[3])+"DGMB'!D"+str(O_O),cell_format)
                        rowOO+=1
                colOP+=1
                colOO+=1


            elif ws['A28'].value == None:        
                for O_P in range(6,30):
                    if O_P==6 or O_P ==7:
                        excel_file_name.write_formula(rowOP,colOP, "='"+str(MB_name[3])+"DGMB'!B"+str(O_P),cell_format)
                        rowOP+=1
                    elif O_P ==8:
                        excel_file_name.write_formula(rowOP,colOP, '0',cell_format)
                        rowOP+=1
                    elif O_P>=9:
                        excel_file_name.write_formula(rowOP,colOP, "='"+str(MB_name[3])+"DGMB'!B"+str(O_P-1),cell_format)
                        rowOP+=1
                for O_O in range(6,30):
                    cell_cord = 'C'+ str(O_O-2)
                    #O_O_sum += int(ws[cell_cord].value)
                    if O_O==6 or O_O ==7:
                        excel_file_name.write_formula(rowOO,colOO, "='"+str(MB_name[3])+"DGMB'!D"+str(O_O),cell_format)
                        rowOO+=1
                    elif O_O ==8:
                        excel_file_name.write_formula(rowOO,colOO, '0',cell_format)
                        rowOO+=1
                    elif O_O>=9:
                        excel_file_name.write_formula(rowOO,colOO, "='"+str(MB_name[3])+"DGMB'!D"+str(O_O-1),cell_format)
                        rowOO+=1
                    #print(O_O_sum)
                    #excel_file_name.write_formula(rowOO,colOO, "='"+str(MB_name[3])+"DGMB'!D"+str(O_O),cell_format)
                colOP+=1
                colOO+=1

            else:                    
                for O_P in range(6,30):
                    excel_file_name.write_formula(rowOP,colOP, "='"+str(MB_name[3])+"DGMB'!B"+str(O_P),cell_format)
                    rowOP+=1
                for O_O in range(6,30):
                    cell_cord = 'C'+ str(O_O-1)
                    #O_O_sum += int(ws[cell_cord].value)
                    #print(O_O_sum)
                    excel_file_name.write_formula(rowOO,colOO, "='"+str(MB_name[3])+"DGMB'!D"+str(O_O),cell_format)
                    rowOO+=1
                colOP+=1
                colOO+=1
        

def import_DDG(starting_column,starting_row,row_range_start, row_range_end, col_range_start, col_range_end, excel_file_name, sheet_name, plik_excel, file_flag):
    cell_format=plik_excel.add_format({'font_size':'11', 'border':True,'align':'center'})
    colDGP=starting_column
    for row in range(row_range_start,row_range_end):  
        RDG_columns=[]

        for col in range(col_range_start,col_range_end):
            if file_flag == True:
                RDG_columns.append(sheet_name.cell(row=row,column=col).value)
            else:
                RDG_columns.append(0)

        rowDGP=starting_row
							
        for valueDGP in RDG_columns:
            excel_file_name.write(rowDGP,colDGP, valueDGP, cell_format)
           
            rowDGP+=1
        colDGP+=1

def file_loop(ws,start_col, start_row, days_range):
    #month_volumen=0
    energy_list=[]
    for col in range(start_col,days_range+start_col):
        day_volumen=0
        day_vol_list=[]
        
        for row in range(start_row,start_row+24):
            try:
                x=int(ws.cell(row=row,column=col).value)
                day_vol_list.append(x)
                day_volumen+=x

            except TypeError:
                
                day_vol_list.append(0)
                day_volumen+=0

            

        energy_list.append(day_vol_list)
    return energy_list

def sub_data(plik_excel,days_range):
    wb1=load_workbook(plik_excel, data_only = True)            
    ws=wb1.active
    consumption=file_loop(ws,2,32,days_range)
    production=file_loop(ws,34,32,days_range)
       
    return consumption,production


def data_subtraction(days_range, excel_file_name, plik_excel):
    cell_format_ok=plik_excel.add_format({'font_size':'11','align':'center', 'bg_color':'green'})
    cell_format_error=plik_excel.add_format({'font_size':'11','align':'center', 'bg_color':'red'})
    cell_format_check=plik_excel.add_format({'font_size':'12','align':'center', 'bg_color':'yellow', 'bold':True})

    excel_file_name.conditional_format(59,1,82,1+days_range-1, {'type':'cell', 'criteria':'==', 'value':'0', 'format':cell_format_ok})
    excel_file_name.conditional_format(59,33,82,33+days_range-1, {'type':'cell', 'criteria':'==', 'value':'0', 'format':cell_format_ok})
    excel_file_name.conditional_format(59,1,82,1+days_range-1, {'type':'cell', 'criteria':'!=', 'value':'0', 'format':cell_format_error})
    excel_file_name.conditional_format(59,33,82,33+days_range-1, {'type':'cell', 'criteria':'!=', 'value':'0', 'format':cell_format_error})
    
    alpha=['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z','AA','AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM','AN','AO','AP','AQ','AR','AS','AT','AU','AV','AW','AX','AY','AZ','BA','BB','BC','BD','BE','BF','BG','BH','BI','BJ','BK','BL','BM','BN','BO','BP','BQ','BR','BS','BT','BU','BV','BW']
    add_value_wire=4
    add_value_rdg=31
    row_add=58

    excel_file_name.write_formula('AG56','=SUM(B60:BL83)',cell_format_check)
    excel_file_name.write_formula('AG57','=MIN(B60:BL83)',cell_format_check)
    excel_file_name.write_formula('AG58','=MAX(B60:BL83)',cell_format_check)

    for hours in range(24):
        add_value_wire+=1
        add_value_rdg+=1
        alpha_letter=1
        row_add+=1
                        
        col_add=1
        alpha_letter=1
        for days in range(days_range):
            excel_file_name.write_formula(row_add,col_add, "="+str(alpha[alpha_letter])+str(add_value_wire)+'-'+str(alpha[alpha_letter])+str(add_value_rdg))
            col_add+=1
            alpha_letter+=1

        col_add=33
        alpha_letter=33    
        for days in range(days_range):
            excel_file_name.write_formula(row_add,col_add, "="+str(alpha[alpha_letter])+str(add_value_wire)+'-'+str(alpha[alpha_letter])+str(add_value_rdg))
            col_add+=1
            alpha_letter+=1
         
def open_xls_as_xlsx(filename):
    # first open using xlrd
    book = xlrd.open_workbook(filename)
    index = 0
    nrows, ncols = 0, 0
    while nrows * ncols == 0:
        sheet = book.sheet_by_index(index)
        nrows = sheet.nrows
        ncols = sheet.ncols
        index += 1

    # prepare a xlsx sheet
    book1 = Workbook()
    sheet1 = book1.get_active_sheet()

    for row in range(1, nrows):
        for col in range(1, ncols):
            sheet1.cell(row=row, column=col).value = sheet.cell_value(row, col)

    #book1.save('asd.xlsx')
    return book1 


def add_chart(plik_excel,excel_file_name, first_col,first_row,days_range, chart_place):
    alpha=['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z','AA','AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM','AN','AO','AP','AQ','AR','AS','AT','AU','AV','AW','AX','AY','AZ','BA','BB','BC','BD','BE','BF','BG','BH','BI','BJ','BK','BL','BM','BN','BO','BP','BQ','BR','BS','BT','BU','BV','BW']
    
    chart=plik_excel.add_chart({'type':'line'})
    for n in range(days_range):
        chart_values='=sprawdzenie!$'+str(alpha[first_col-1+n])+'$'+str(first_row)+':$'+str(alpha[first_col-1+n])+'$'+str(first_row+23)
        chart.add_series({'values':chart_values, 'data_labels':{'value':False}})


    chart.set_size({'x_scale': 4, 'y_scale': 3})
    excel_file_name.insert_chart(chart_place, chart)

def change_label_text(variable, text):
    variable.set(text)
    

def raport_summary(item):
    res1=[sum(x) for x in item[0]]
    sum_res_1=sum(res1)
    res2=[sum(x) for x in item[1]]
    sum_res_2=sum(res2)
    return sum_res_1, sum_res_2


def extrems(list_one,list_two,days_range):
    z=[]
    sum_value=0
    lst1 = [item for item in list_one]
    lst2 = [item for item in list_two]
    for k in range(2):
        #sum_list=[]
        for d in range(0,days_range):
            
            for h in range(24):
                f=lst1[k][d][h]
                b=lst2[k][d][h]
                z.append(f-b)
    sum_value=sum(z)  
    min_value=min(z)
    max_value=max(z)

    return sum_value,max_value, min_value

def report_chart_sheet(plik_excel,excel_file_name,list_one,days_range, MB_name,odstep_y):

    chart_OP=plik_excel.add_chart({'type':'line'})
    chart_OO=plik_excel.add_chart({'type':'line'})

    lst1 = [item for item in list_one]
    for d in range(days_range):
        chart_values = lst1[odstep_y][0][d]
        chart_value_str1=str(chart_values).replace(".0,",";")
        chart_value_str2=str(chart_value_str1).replace(".0","")
        chart_value_str3=str(chart_value_str2).replace("]","}")
        chart_value_str4=str(chart_value_str3).replace("[","{")
        chart_value_str5=str(chart_value_str4).replace(" ","")
        chart_OP.add_series({'values':chart_value_str5})

    for d in range(days_range):
        chart_values_abc = lst1[odstep_y][1][d]
        chart_value_stra=str(chart_values_abc).replace(".0,",";")
        chart_value_strb=str(chart_value_stra).replace(".0","")
        chart_value_strc=str(chart_value_strb).replace("]","}")
        chart_value_strd=str(chart_value_strc).replace("[","{")
        chart_value_stre=str(chart_value_strd).replace(" ","")
        chart_OO.add_series({'values':chart_value_stre})

    chart_OP.set_size({'x_scale': 3, 'y_scale': 2})
    chart_OP.set_title({'name': MB_name+' POBÓR'})
    excel_file_name.insert_chart('B'+str(3+int(odstep_y)*33), chart_OP)

    chart_OO.set_size({'x_scale': 3, 'y_scale': 2})
    chart_OO.set_title({'name': MB_name+' ODDANIE' })
    excel_file_name.insert_chart('Z'+str(3+int(odstep_y)*33), chart_OO)



def report_file(location,start_date,end_date,report_data,WIRE,POB_name,to_char_data):
    report_date=datetime.now()
    report_excel = xlsxwriter.Workbook(str(location)+'/'+'_Raport sprawdzania POB za '+datetime.strftime(start_date,'%d-%m-%Y')+"-"+datetime.strftime(end_date,'%d-%m-%Y')+', wykonano '+datetime.strftime(report_date,"%Y-%m-%d_%H-%M-%S")+'.xlsx')    
    report_sheet=report_excel.add_worksheet('Raport')
    chart_sheet=report_excel.add_worksheet('Wykresy')
    chart_sheet.set_zoom(60)
    report_sheet.set_column(0,5,22.5)
    cell_format_standard=report_excel.add_format({'font_size':'11','align':'center'})
    cell_format_error=report_excel.add_format({'font_size':'11','align':'center', 'bg_color':'red'})
    report_sheet.conditional_format(3,3,3+len(report_data),5, {'type':'cell', 'criteria':'!=', 'value':'0', 'format':cell_format_error})
   
    report_sheet.write('A1', 'Data wykonania raportu:',cell_format_standard)
    report_sheet.write('B1', datetime.strftime(report_date,"%Y-%m-%d_%H-%M-%S"),cell_format_standard)
    report_sheet.write('A3', 'Kod MB',cell_format_standard)
    report_sheet.write('B3', 'Suma Poboru (WIRE)',cell_format_standard)
    report_sheet.write('C3', 'Suma Oddania (WIRE)',cell_format_standard)
    report_sheet.write('D3', 'ΔE Suma (WIRE-RDG):',cell_format_standard)
    report_sheet.write('E3', 'ΔE Pobór (WIRE-RDG):',cell_format_standard)
    report_sheet.write('F3', 'ΔE Oddanie (WIRE-RDG)',cell_format_standard)

    row=3
    for n in range(len(report_data)):
        report_sheet.write(row,0,report_data[n][0],cell_format_standard)
        report_sheet.write(row,1,report_data[n][1],cell_format_standard)
        report_sheet.write(row,2,report_data[n][2],cell_format_standard)
        report_sheet.write(row,3,report_data[n][3],cell_format_standard)
        report_sheet.write(row,4,report_data[n][4],cell_format_standard)
        report_sheet.write(row,5,report_data[n][5],cell_format_standard)
        row+=1
    
    for x in range(len(report_data)):
        report_chart_sheet(report_excel, chart_sheet, to_char_data,(end_date-start_date).days+1, report_data[x][0], x)


    report_excel.close()
