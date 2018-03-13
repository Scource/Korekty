'''
Do zrobienia:
trye/expecty/zabezpieczenia/pop-upy
'''

from tkinter import Tk, Label, Button, Entry, StringVar, filedialog, Listbox, messagebox
from tkinter.filedialog import askopenfilename
from openpyxl import load_workbook, Workbook
from datetime import datetime, timedelta, date
import xlsxwriter
import string
from korekty_def_file_v6 import *
import sys




class MainWindow:

    
    
    def __init__(self, master):
        self.master=master
        master.title("porównywanie Edziowym sposobem")
             
        
        #self.frame1=Frame(master)

        self.button_skok=Button(master, text="Importuj dane z WIRE", command=lambda: self.file_choose("skok"))
        self.button_skok.pack()

        self.label_WIRE_text = StringVar()
        self.label_wire_files=Label(master, textvariable=self.label_WIRE_text)
        self.label_wire_files.pack()

        self.button_skome=Button(master, text="Importuj dane z RDG", command=lambda: self.file_choose("skome"))
        self.button_skome.pack()

        
        self.label_RDG_text = StringVar()
        self.label_rdg_files=Label(master, textvariable=self.label_RDG_text)
        self.label_rdg_files.pack()

        self.button_direcotory=Button(master, text="wybierz lokację do zapisu plików", command=lambda: self.choose_save_location())
        self.button_direcotory.pack()
        
        self.button_start=Button(master, text="Start", command=lambda: self.file_create(RDG_list,WIRE_list))
        self.button_start.pack()


        
    def file_create(self, RDG, WIRE):

        count_list_RDG=set([POB[1] for POB in RDG])
        count_list_WIRE=set([POB[1] for POB in WIRE])
        time_delta_check=set([(POB[3]-POB[2]).days+1 for POB in RDG])
        timelapse=set([POB[2] for POB in RDG])
        #timelapse1=datetime.strptime(str(set([POB[2] for POB in RDG])),'%Y,%m,%d,%H,%M')
        timelapse2=set([POB[3] for POB in RDG])
        #timelapse2=datetime.strptime(str(set([POB[3] for POB in RDG])),'%Y-%m-%d')

        if count_list_RDG==count_list_WIRE:
            messagebox.showinfo("","Zaimportowano  " + str(len(count_list_RDG)) + "  POB\n"
                                "Rozpatrywany okres to\n" 
                                + str(timelapse) +" do "+str(timelapse2))
        else:
            messagebox.showerror("Błąd","Something is no yes, try again!")
            return

        global index
        POB_was=[]
        report=[]
        to_char_data=[]
        no_files_list=[]
        for POB_name in RDG:
            
            if POB_name[1] not in POB_was:
                POB_was.append(POB_name[1])
                
                plik_excel = xlsxwriter.Workbook(str(location)+'/'+POB_name[1]+'.xlsx')
                
                sprawdzenie=plik_excel.add_worksheet('sprawdzenie')
                sprawdzenie.set_zoom(55)
  
                import_WIRE_file(WIRE,POB_name, 33, 1, 4, 4, sprawdzenie, plik_excel)
                data_subtraction((POB_name[3]-POB_name[2]).days+1, sprawdzenie, plik_excel)
                
                
                add_chart(plik_excel, sprawdzenie, 2, 5, (POB_name[3]-POB_name[2]).days+1,'B86')
                add_chart(plik_excel, sprawdzenie, 34, 5, (POB_name[3]-POB_name[2]).days+1,'AH86')

                if POB_name[0][19:] == "_P":
                    
                    new_xlsx=open_xls_as_xlsx(POB_name[4])
                    wb2=new_xlsx
                    ws2=wb2.active
                    import_DDG(1,31,7,(POB_name[3]-POB_name[2]).days+8, 1, 25, sprawdzenie, ws2, plik_excel, True)

                    match=[item for item in RDG if item[1] == POB_name[1] and item[0] != POB_name[0]]
                    if len(match)==1:                    
                        #index = RDG.index(item)
                        new_xlsx=open_xls_as_xlsx(RDG[index][4])
                        wb2=new_xlsx
                        ws2=wb2.active
                        import_DDG(33,31,7, (RDG[index][3]-RDG[index][2]).days+8, 1, 25, sprawdzenie, ws2, plik_excel, True)
                            
                    elif len(match)==0:
                        index = RDG.index(POB_name)
                        import_DDG(33,31,7, (RDG[index][3]-RDG[index][2]).days+8, 1, 25, sprawdzenie, ws2, plik_excel, False)
                        if str(POB_name[1])+"_O" not in no_files_list:
                            no_files_list.append(str(POB_name[1])+"_O")
                    elif len(match)>1:
                        messagebox.showerror("Błąd", "Dwa pliki RDG z tym samym kodem MB")
                        sys.exit(1)
  
                else:
                    new_xlsx=open_xls_as_xlsx(POB_name[4])
                    wb2=new_xlsx
                    ws2=wb2.active
                    import_DDG(33,31,7,(POB_name[3]-POB_name[2]).days+8, 1, 25, sprawdzenie, ws2, plik_excel, True)

                    
                    match=[item for item in RDG if item[1] == POB_name[1] and item[0] != POB_name[0]]
                    if len(match)==1:
                            index = RDG.index(match[0])
                            new_xlsx=open_xls_as_xlsx(RDG[index][4])
                            wb2=new_xlsx
                            ws2=wb2.active
                            import_DDG(1,31,7, (RDG[index][3]-RDG[index][2]).days+8, 1, 25, sprawdzenie, ws2, plik_excel, True)
                    elif len(match)==0:
                        index = RDG.index(POB_name)
                        import_DDG(1,31,7, (RDG[index][3]-RDG[index][2]).days+8, 1, 25, sprawdzenie, ws2, plik_excel, False)
                        if str(POB_name[1])+"_P" not in no_files_list:
                            no_files_list.append(str(POB_name[1])+"_P")
                    elif len(match)>1:
                        messagebox.showerror("Błąd", "Dwa pliki RDG z tym samym kodem MB")
                        sys.exit(1)

                plik_excel.close()

                cons_val_file=raport_summary(wire_volumen_sum(WIRE,POB_name))[0]
                prod_val_file=raport_summary(wire_volumen_sum(WIRE,POB_name))[1]
                dupa33=wire_volumen_sum(WIRE,POB_name)
                sum_val_file=extrems(wire_volumen_sum(WIRE,POB_name),sub_data(str(location)+'/'+POB_name[1]+'.xlsx',(POB_name[3]-POB_name[2]).days+1),(POB_name[3]-POB_name[2]).days+1)[0]
                max_val_file=extrems(wire_volumen_sum(WIRE,POB_name),sub_data(str(location)+'/'+POB_name[1]+'.xlsx',(POB_name[3]-POB_name[2]).days+1),(POB_name[3]-POB_name[2]).days+1)[1]
                min_val_file=extrems(wire_volumen_sum(WIRE,POB_name),sub_data(str(location)+'/'+POB_name[1]+'.xlsx',(POB_name[3]-POB_name[2]).days+1),(POB_name[3]-POB_name[2]).days+1)[2]
                file_tuple=(POB_name[1],cons_val_file,prod_val_file,sum_val_file, max_val_file,min_val_file)
                
                report.append(file_tuple)                
                to_char_data.append(dupa33)

        report_file(location,POB_name[2],POB_name[3],report,WIRE,POB_name,to_char_data)

        if len(no_files_list)==0:
            messagebox.showinfo("Koniec", "Zadanie zakończone")
        else:
            message_test =""
            for a in no_files_list:
                message_test = message_test+str(a)+"\n"
            messagebox.showinfo("Koniec", "MB bez plików - zera przyjęto dla:\n"+message_test)
                
    def file_choose(self, file_dir):
        if file_dir == "skok":
            wire_files = []
            wire_dir = filedialog.askopenfilenames()
            wire_files = list(wire_dir)
            WIRE_string=[]
            for i in wire_files:
                WIRE_tuple=(datetime.strptime(i[-14:-4],'%Y-%m-%d'),i[-34:-15],i[:-40],i[:-39]+'['+i[-39:]+']',i)
                WIRE_string.append(WIRE_tuple)
            global WIRE_list
            
            WIRE_list = WIRE_string
            change_label_text(self.label_WIRE_text, str(len(WIRE_list)))
            
                    
        elif file_dir == "skome":
            skome_files = []
            skome_dir = filedialog.askopenfilenames()
            skome_files = list(skome_dir)
            RDG_files=[]
            for i in skome_files:
                wb1=load_workbook(i)            
                ws=wb1.active
                x = str(ws['D3'].value)
                y = str(ws['D4'].value)
                if "***" in x:
                    RDG_tuple=(x[3:24],x[3:22],datetime.strptime(y[0:10],'%d-%m-%Y'),datetime.strptime(y[19:29],'%d-%m-%Y')-timedelta(days=1),i)
                else:
                    RDG_tuple=(x[0:21],x[0:19],datetime.strptime(y[0:10],'%d-%m-%Y'),datetime.strptime(y[19:29],'%d-%m-%Y')-timedelta(days=1),i)
                RDG_files.append(RDG_tuple)
            global RDG_list
            
            RDG_list=RDG_files
            change_label_text(self.label_RDG_text, str(len(RDG_list)))
            #print(RDG_list)

    def choose_save_location(self):
        global location
        location=filedialog.askdirectory()

root = Tk()
my_gui = MainWindow(root)
root.mainloop()

